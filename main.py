import json
import os
import sys
from filename import create_summary_file_path, get_day_from_file_name
from openpyxl import load_workbook
from shift_change import process_excel
from summary import create_summary_file, transfer_data
from access import check_access

if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: exauto <excel_file_name>")
        exit(1)
    elif not os.path.exists(sys.argv[1]):
        print(f"{sys.argv[1]} does not exist!")
        exit(1)
    else:
        excel_file_name = sys.argv[1]
        process_excel(excel_file_name)

        access = check_access()

        if not access:
            print("Could not create summary file.")
            sys.exit(1)

        summary_file_path = create_summary_file_path(sys.argv[1])

        # Retrieve summary map
        script_dir = os.path.dirname(os.path.abspath(__file__))
        workbook_map_file_path = os.path.join(
            script_dir, "summary_workbook_map.json")

        with open(workbook_map_file_path, "r") as summary_workbook_map_file:
            summary_map = json.load(summary_workbook_map_file)

        # Create summary file if it doesn't exist
        if not os.path.exists(summary_file_path):
            create_summary_file(sys.argv[1], summary_map)

        # open the files you need, and get the date_number
        day_number = get_day_from_file_name(sys.argv[1])

        try:
            source_workbook = load_workbook(sys.argv[1], data_only=True)
            summary_workbook = load_workbook(summary_file_path)
        except FileNotFoundError as e:
            print(f"Error: {e}")
            sys.exit(1)

        transfer_data(source_workbook, summary_workbook,
                      summary_map, day_number)

        # Close files
        source_workbook.close()
        summary_workbook.save(summary_file_path)
        summary_workbook.close()
