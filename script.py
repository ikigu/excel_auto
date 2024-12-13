#!/usr/bin/python

from filename import create_new_file_path
from dotenv import load_dotenv
from openpyxl import load_workbook
import json
import os
import shutil
import sys


def process_excel(old_file_name):
    """
    Process an Excel file to copy values (not formulas)
    from closing stock columns/rows to opening stock columns/rows

    Args:
        file_name (str): The name of the Excel file to process.
    """

    load_dotenv()
    debug_mode = os.getenv("DEBUG") == "True"

    # Create today's copy of the Excel workbook to edit
    updated_file_name = create_new_file_path(old_file_name)
    shutil.copy(old_file_name, updated_file_name)
    print(f"{updated_file_name} has been created.\nCopying closing stock to opening stock...")

    # Retrieve workbook map
    script_dir = os.path.dirname(os.path.abspath(__file__))
    workbook_map_file_path = os.path.join(script_dir, "workbook_map.json")

    try:
        with open(workbook_map_file_path, "r") as workbook_map_file:
            workbook_map = json.load(workbook_map_file)
    except FileNotFoundError:
        print(f"{workbook_map_file_path} not found. File not copied!")
        os.remove(updated_file_name)
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"Error: Failed to decode workbook_map.json. Details: {e}")
        os.remove(updated_file_name)
        sys.exit(1)

    # Load original (read-only) and updated (editable) workbooks
    try:
        original_workbook = load_workbook(old_file_name, data_only=True)
        updated_workbook = load_workbook(updated_file_name)
    except FileNotFoundError:
        print(f"Error: File '{old_file_name}' not found.")
        os.remove(updated_file_name)
        sys.exit(1)

    # Run through each worksheet
    for sheet in workbook_map:
        sheet_name: str = sheet["sheet_name"]
        column_map: dict = sheet["column_map"]
        row_map: dict = sheet["row_map"]
        first_data_column: str = sheet["first_data_column"]
        final_data_column: str = sheet["final_data_column"]
        first_data_row: int = sheet["first_data_row"]
        rows_to_ignore: list[int] = sheet["rows_to_ignore"]
        columns_to_clear: list[str] = sheet["columns_to_clear"]
        rows_to_clear: list[int] = sheet["rows_to_clear"]

        # Check that sheet exists
        if sheet_name not in original_workbook.sheetnames:
            print(f"Error: {sheet_name} doesn't exist. File not copied!")
            os.remove(updated_file_name)
            sys.exit(1)

        old_sheet = original_workbook[sheet_name]
        new_sheet = updated_workbook[sheet_name]

        if debug_mode:
            print(f"{sheet_name}\n\n-----------\n\n")

        # Copy columns: closing stock --> opening stock
        for closing_stock_column, opening_stock_column in column_map.items():
            for row_number in range(first_data_row, new_sheet.max_row + 1):

                if row_number in rows_to_ignore:
                    continue

                closing_cell = old_sheet[f"{closing_stock_column}{row_number}"]
                opening_cell = new_sheet[f"{opening_stock_column}{row_number}"]

                if isinstance(closing_cell.value, (int, float)):
                    opening_cell.value = closing_cell.value

                    if debug_mode:
                        print(
                            f"Copied data from {closing_cell} --> {opening_cell}")

        # Clear columns that affect closing stock, except opening column
        for column_to_clear in columns_to_clear:
            for row_number in range(first_data_row, new_sheet.max_row + 1):
                cell_to_clear = new_sheet[f"{column_to_clear}{row_number}"]

                if isinstance(cell_to_clear.value, (int, float)):
                    cell_to_clear.value = None

                    if debug_mode:
                        print(f"Cleared data from {cell_to_clear}")

        # Copy rows: closing stock --> opening stock
        for closing_stock_row, opening_stock_row in row_map.items():
            for column in range(ord(first_data_column), ord(final_data_column) + 1):

                closing_cell = old_sheet[f"{chr(column)}{int(closing_stock_row)}"]
                opening_cell = new_sheet[f"{chr(column)}{opening_stock_row}"]

                if isinstance(closing_cell.value, (int, float)):
                    opening_cell.value = closing_cell.value

                    if debug_mode:
                        print(
                            f"Copied data from {closing_cell} --> {opening_cell}")

        # Clear rows that affect closing stock except opening stoc
        for row_to_clear in rows_to_clear:
            for column in range(ord(first_data_column), ord(final_data_column) + 1):
                cell_to_clear = f"{chr(column_to_clear)}{row_to_clear}"

                if isinstance(cell_to_clear.value, (int, float)):
                    cell_to_clear.value = None

                    if debug_mode:
                        print(f"Cleared data from {cell_to_clear}")

        print(f"Data from {sheet_name} has been copied successfully.")

    # Save the updated workbook
    updated_workbook.save(updated_file_name)
    # Close workbooks
    original_workbook.close()
    updated_workbook.close()
    print(f"Workbook updated and saved as '{updated_file_name}'.")

# Entry point for the program


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print("Usage: python script_name.py <excel_file_name>")
    else:
        excel_file_name = sys.argv[1]
        process_excel(excel_file_name)
